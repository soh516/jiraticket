import os
import glob
import shutil
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd

month = datetime.now().strftime('%h')
year = datetime.now().strftime('%Y')
outputFileName = 'rc_' + month + year + '.xlsx'

# openpyxl is python library to read/write excel

dest_wb = Workbook()

# dir where my excel files are. The combined file will also be saved here. 
dir_containing_files = "/home/soh516/Documents"

# os.walk() returns current_path, directories in current_path, files in current_path
for root, dirs, filenames in os.walk(dir_containing_files):
    for file in filenames:
        if file.endswith('.csv'):
            # file is with extension. Get rid of extension as file_name
            csv_file_path = os.path.abspath(os.path.join(root, file))
            # pay attention to the delimiter here. We assume it is "," not ";"
            read_file = pd.read_csv(csv_file_path, delimiter=',')
            file_name = file.split('.')[0]
            excel_file_name = file_name + '.xlsx'
            file_path = os.path.abspath(os.path.join(root, excel_file_name))
            read_file.to_excel(file_path, index=False, header=True)

            dest_wb.create_sheet(file_name)
            dest_ws = dest_wb[file_name]

            source_wb = load_workbook(file_path)
            # This assumes there is only one tab
            source_sheet = source_wb.active
            for row in source_sheet.rows:
                for cell in row:
                    dest_ws[cell.coordinate] = cell.value

            os.remove(file_path)

outputFileName = dir_containing_files + '/' + outputFileName
dest_wb.remove(dest_wb['Sheet'])
dest_wb.save(outputFileName)

# copy all useful files to my onedrive
fullMonth = datetime.now().strftime('%B')
finalSaveDir = "/home/soh516/onedrive-usask" + "/jira/" + fullMonth
isExist = os.path.exists(finalSaveDir)

if not isExist:
    os.makedirs(finalSaveDir)

for root, dirs, filenames in os.walk(dir_containing_files):
    for file in filenames:
        src_file_path = os.path.abspath(os.path.join(root, file))
        shutil.copy(src_file_path, finalSaveDir + '/')
